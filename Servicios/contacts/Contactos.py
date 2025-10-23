# Servicios/contacts/Contactos.py
import socket
import json
import threading
import logging
import time
import os
import re
import csv
import base64
from io import BytesIO, StringIO
from datetime import datetime, date
from typing import Any, Dict, List, Optional

from pymongo import MongoClient, ASCENDING
from pymongo.errors import PyMongoError, DuplicateKeyError
from bson.objectid import ObjectId

try:
    from openpyxl import load_workbook, Workbook
except Exception:
    load_workbook = None
    Workbook = None
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - [CONTACTOS] - %(levelname)s - %(message)s'
)

def _jsonline(obj: dict) -> bytes:
    return (json.dumps(obj, ensure_ascii=False) + "\n").encode("utf-8")

def _safe_object_id(s: str) -> Optional[ObjectId]:
    try:
        return ObjectId(s)
    except Exception:
        return None

def _norm_phone(s: str) -> str:
    digits = re.sub(r"\D+", "", s or "")
    if not digits:
        return ""
    if digits.startswith("56"):
        digits = digits[2:]
    if len(digits) == 8 and not digits.startswith("9"):
        digits = "9" + digits
    if len(digits) > 9:
        digits = digits[-9:]
    if len(digits) < 8:
        return "+56" + digits if digits else ""
    return "+56" + digits

def _digits_regex(term: str) -> Optional[re.Pattern]:
    digits = re.sub(r"\D+", "", term or "")
    if not digits:
        return None
    pattern = r"\D*".join(map(re.escape, digits))
    try:
        return re.compile(pattern, re.IGNORECASE)
    except re.error:
        return None

def _serialize(doc: Dict[str, Any]) -> Dict[str, Any]:
    """
    Convierte un doc Mongo a JSON serializable.
    Además: normaliza 'telefono' (string o lista) para que SIEMPRE salga +569XXXXXXXX.
    """
    if not doc:
        return {}
    out: Dict[str, Any] = {}
    for k, v in doc.items():
        if k == "_id":
            out["id"] = str(v)
            continue
        if k == "telefono":
            if isinstance(v, list):
                tel_list = []
                for x in v:
                    n = _norm_phone(str(x))
                    if n:
                        tel_list.append(n)
                out["telefono"] = tel_list
            else:
                out["telefono"] = _norm_phone(str(v)) if (str(v).strip()) else ""
            continue
        if isinstance(v, (datetime, date)):
            out[k] = v.isoformat()
        elif isinstance(v, ObjectId):
            out[k] = str(v)
        else:
            out[k] = v
    return out

class ContactosService:
    def __init__(
        self,
        client_id: str = os.getenv("CLIENT_ID", "contactos_service"),
        bus_host: str = os.getenv("BUS_HOST", "bus"),
        bus_port: int = int(os.getenv("BUS_PORT", "5000")),
        mongo_uri: str = os.getenv("MONGO_URI", "mongodb://app_user:app_password_123@mongodb:27017/?authSource=admin"),
        mongo_db: str = os.getenv("MONGO_DB", "usuarios_db"),
        mongo_coll: str = os.getenv("MONGO_COLL", "contactos")
    ):
        self.client_id = client_id
        self.bus_host = bus_host
        self.bus_port = bus_port
        self.socket: Optional[socket.socket] = None
        self.connected = False
        self.running = False

        self.mongo_uri = mongo_uri
        self.mongo_db = mongo_db
        self.mongo_coll = mongo_coll
        self.db_client: Optional[MongoClient] = None
        self.col = None

    def init_db(self) -> bool:
        try:
            self.db_client = MongoClient(self.mongo_uri, serverSelectionTimeoutMS=6000)
            self.db_client.admin.command("ping")
            db = self.db_client[self.mongo_db]
            self.col = db[self.mongo_coll]
            self.col.create_index([("departamento", ASCENDING), ("nombre", ASCENDING)])
            logging.info("MongoDB listo (Contactos)")
            return True
        except PyMongoError as e:
            logging.error(f"Error conectando a MongoDB: {e}")
            return False

    def connect_bus(self) -> bool:
        try:
            self.socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            logging.info(f"Conectando al BUS en {self.bus_host}:{self.bus_port}")
            self.socket.connect((self.bus_host, self.bus_port))

            register_message = {
                "type": "REGISTER",
                "client_id": self.client_id,
                "kind": "service",
                "service": "Contactos",
            }
            self.socket.sendall(_jsonline(register_message))

            self.socket.settimeout(2.0)
            buf = ""
            try:
                while "\n" not in buf:
                    chunk = self.socket.recv(4096).decode("utf-8")
                    if not chunk:
                        break
                    buf += chunk
            finally:
                self.socket.settimeout(None)

            if "\n" in buf:
                line, _rest = buf.split("\n", 1)
                line = line.strip()
                if line:
                    ack = json.loads(line)
                    if ack.get("type") == "REGISTER_ACK" and ack.get("status") == "success":
                        self.connected = True
                        self.running = True
                        threading.Thread(target=self._listen_messages, daemon=True).start()
                        logging.info("Registrado en BUS como servicio 'Contactos'")
                        logging.info("Servicio de Contactos iniciado correctamente")
                        return True

            logging.error(f"Error en registro BUS: {buf!r}")
            return False

        except Exception as e:
            logging.error(f"Error conectando al BUS: {e}")
            return False

    def _listen_messages(self):
        logging.info("Iniciando escucha de mensajes del BUS...")
        buf = ""
        while self.running and self.connected:
            try:
                data = self.socket.recv(65536)
                if not data:
                    logging.warning("⚠️ Conexión cerrada por el BUS")
                    self.connected = False
                    break

                buf += data.decode("utf-8")
                while "\n" in buf:
                    line, buf = buf.split("\n", 1)
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        message = json.loads(line)
                        self._handle_message(message)
                    except json.JSONDecodeError as e:
                        logging.error(f"Error decodificando JSON NDJSON: {e}")

            except Exception as e:
                if self.running:
                    logging.error(f"Error recibiendo: {e}")
                    self.connected = False
                break
        logging.info("Listener detenido")

    def _send(self, d: Dict[str, Any]) -> bool:
        if not self.connected:
            logging.error("No conectado al BUS. No se puede enviar mensaje.")
            return False
        try:
            d.setdefault("sender", self.client_id)
            self.socket.sendall(_jsonline(d))
            logging.info(f"Mensaje enviado - Tipo: {d.get('type')}")
            return True
        except Exception as e:
            logging.error(f"Error enviando mensaje: {e}")
            self.connected = False
            return False

    def _respond_direct(self, target: str, payload: Dict[str, Any], correlation_id: Optional[str] = None):
        out: Dict[str, Any] = {
            "type": "DIRECT",
            "target": target,
            "payload": payload,
            "service": "Contactos",
        }
        if correlation_id:
            out["header"] = {"correlationId": correlation_id}
        self._send(out)

    def send_broadcast(self, payload: Dict[str, Any]) -> bool:
        return self._send({"type": "BROADCAST", "payload": payload})

    def _handle_message(self, message: Dict[str, Any]):
        mtype = message.get("type")
        sender = message.get("sender", "UNKNOWN")
        logging.info(f"Mensaje recibido - Tipo: {mtype}, De: {sender}")

        if mtype == "REQUEST":
            header = message.get("header") or {}
            action = (header.get("action") or (message.get("payload") or {}).get("action") or "").strip()
            payload = message.get("payload", {}) or {}
            corr = header.get("correlationId")

            try:
                resp_payload = self._handle_request(action, payload)
                envelope = resp_payload
            except Exception as e:
                logging.exception("Fallo manejando request (no controlado)")
                envelope = {"ok": False, "error": str(e)}

            self._respond_direct(sender, envelope, correlation_id=corr)
            return

        if mtype == "DIRECT":
            logging.info(f"📧 DIRECT de {sender}: {message.get('payload', {})}")
            return

        if mtype == "BROADCAST":
            logging.info(f"BROADCAST de {sender}: {message.get('payload', {})}")
            return

        if mtype == "DELIVERY_ACK":
            logging.info(f"Mensaje entregado a {message.get('target')}")
            return

        if mtype == "ERROR":
            logging.error(f"Error del BUS: {message.get('message')}")
            return
        logging.info(f"REQUEST action={action!r} payload_keys={list(payload.keys())}")

        logging.debug(f"(ignorado) {message}")

    def _require(self, obj: Dict[str, Any], keys: List[str]):
        for k in keys:
            if obj.get(k) in (None, ""):
                raise ValueError(f"Falta parámetro requerido: {k}")

    def _handle_request(self, action: str, p: Dict[str, Any]) -> Dict[str, Any]:
        try:
            if action == "create_contact":
                return self._create_contact(p)
            if action == "get_contact":
                return self._get_contact(p)
            if action == "list_contacts":
                return self._list_contacts(p)
            if action == "update_contact":
                return self._update_contact(p)
            if action == "delete_contact":
                return self._delete_contact(p)
            if action == "search":
                return self._search(p)
            if action == "import":
                return self._import_contacts(p)
            if action == "export":
                return self._export_contacts(p)


            return {"ok": False, "error": f"Acción no soportada: {action}"}

        except DuplicateKeyError:
            return {"ok": False, "error": "Contacto duplicado"}
        except PyMongoError as e:
            return {"ok": False, "error": f"MongoError: {e}"}
        except Exception as e:
            logging.exception("Fallo manejando request")
            return {"ok": False, "error": str(e)}

    def _create_contact(self, p: Dict[str, Any]) -> Dict[str, Any]:
        nombre = p.get("nombre") or p.get("name")
        departamento = p.get("departamento")
        telefono = p.get("telefono") or p.get("phone")
        if not (nombre and departamento and telefono):
            raise ValueError("Se requieren: nombre, departamento, telefono")

        tel_norm = _norm_phone(telefono)
        doc: Dict[str, Any] = {
            "nombre": str(nombre),
            "departamento": str(departamento),
            "telefono": tel_norm or str(telefono),
            "createdAt": datetime.utcnow(),
        }
        email = (p.get("email") or "").strip()
        if email:
            doc["email"] = email
        tags = p.get("tags")
        if isinstance(tags, list):
            doc["tags"] = [str(t) for t in tags]

        res = self.col.insert_one(doc)
        out = _serialize({**doc, "_id": res.inserted_id})
        self.send_broadcast({"event": "contacts.created", "contact_id": out["id"]})
        return {"ok": True, "contact": out}

    def _get_contact(self, p: Dict[str, Any]) -> Dict[str, Any]:
        cid = _safe_object_id(p.get("contact_id") or p.get("id") or "")
        if not cid:
            return {"ok": False, "error": "contact_id inválido"}
        doc = self.col.find_one({"_id": cid})
        if not doc:
            return {"ok": False, "error": "No encontrado"}
        return {"ok": True, "contact": _serialize(doc)}

    def _list_contacts(self, p: Dict[str, Any]) -> Dict[str, Any]:
        limit = max(1, min(int(p.get("limit", 50)), 200))
        page = max(1, int(p.get("page", 1)))
        skip = (page - 1) * limit
        cur = (
            self.col.find({})
            .sort([("departamento", ASCENDING), ("nombre", ASCENDING)])
            .skip(skip)
            .limit(limit)
        )
        items = [_serialize(d) for d in cur]
        total = self.col.count_documents({})
        return {"ok": True, "items": items, "page": page, "limit": limit, "total": total}

    def _update_contact(self, p: Dict[str, Any]) -> Dict[str, Any]:
        cid = _safe_object_id(p.get("contact_id") or p.get("id") or "")
        if not cid:
            return {"ok": False, "error": "contact_id inválido"}

        allowed = {"nombre", "departamento", "telefono", "email", "tags"}
        updates_raw = (p.get("updates") or {})
        updates = {k: v for k, v in updates_raw.items() if k in allowed}

        if "telefono" in updates:
            norm = _norm_phone(str(updates["telefono"]))
            if norm:
                updates["telefono"] = norm

        if "email" in updates and not (updates["email"] or "").strip():
            updates.pop("email")

        if "tags" in updates and not isinstance(updates["tags"], list):
            updates.pop("tags")

        if not updates:
            return {"ok": False, "error": "Nada para actualizar"}

        updates["_updatedAt"] = datetime.utcnow()

        res = self.col.find_one_and_update(
            {"_id": cid},
            {"$set": updates},
            return_document=True
        )
        if not res:
            return {"ok": False, "error": "No encontrado"}

        out = _serialize(res)
        self.send_broadcast({"event": "contacts.updated", "contact_id": out["id"]})
        return {"ok": True, "contact": out}

    def _delete_contact(self, p: Dict[str, Any]) -> Dict[str, Any]:
        cid = _safe_object_id(p.get("contact_id") or p.get("id") or "")
        if not cid:
            return {"ok": False, "error": "contact_id inválido"}

        res = self.col.delete_one({"_id": cid})
        if res.deleted_count == 0:
            return {"ok": False, "error": "No encontrado"}

        self.send_broadcast({"event": "contacts.deleted", "contact_id": str(cid)})
        return {"ok": True}

    def _search(self, p: Dict[str, Any]) -> Dict[str, Any]:
        q: Dict[str, Any] = {}
        term = (p.get("q") or "").strip()
        nombre = (p.get("nombre") or "").strip()
        depto = (p.get("departamento") or "").strip()

        ors: List[Dict[str, Any]] = []
        if term:
            rx_any = re.compile(re.escape(term), re.IGNORECASE)
            ors.extend([{"nombre": rx_any}, {"departamento": rx_any}])

            rx_digits = _digits_regex(term)
            if rx_digits:
                ors.append({"telefono": rx_digits})

            norm = _norm_phone(term)
            if norm:
                ors.append({"telefono": norm})

        if nombre:
            ors.append({"nombre": re.compile(re.escape(nombre), re.IGNORECASE)})
        if depto:
            rx_depto = re.compile(rf"^(?:{re.escape(depto)}$|{re.escape(depto)}.*)", re.IGNORECASE)
            ors.append({"departamento": rx_depto})

        if ors:
            q["$or"] = ors

        limit = max(1, min(int(p.get("limit", 50)), 200))
        cur = (
            self.col.find(q)
            .sort([("departamento", ASCENDING), ("nombre", ASCENDING)])
            .limit(limit)
        )
        items = [_serialize(d) for d in cur]
        return {"ok": True, "contacts": items}

    def disconnect(self):
        logging.info("Desconectando del BUS/Mongo...")
        self.running = False
        self.connected = False
        if self.socket is not None:
            try:
                self.socket.close()
            except:
                pass
        if self.db_client is not None:
            try:
                self.db_client.close()
            except:
                pass
        logging.info("Contactos detenido")
        
    def _import_contacts(self, p: Dict[str, Any]) -> Dict[str, Any]:
        fmt = (p.get("format") or "xlsx").strip().lower()
        b64 = (p.get("fileContent") or "").strip()
        if not b64:
            return {"ok": False, "error": "Falta fileContent (base64 del archivo)"}

        inserted = 0
        updated = 0
        skipped = 0
        errors: List[str] = []
        rows: List[List[str]] = []

        # --- leer archivo (igual que tu versión) ---
        try:
            raw = base64.b64decode(b64)
            if fmt == "csv":
                text = raw.decode("utf-8", errors="ignore")
                text = text.replace(";", ",")             # acepta ; o ,
                reader = csv.reader(StringIO(text))
                rows = [list(r) for r in reader]
            else:
                if load_workbook is None:
                    return {"ok": False, "error": "openpyxl no instalado en el contenedor"}
                wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
                ws = wb.worksheets[0]
                for r in range(1, ws.max_row + 1):
                    rows.append([
                        (ws.cell(row=r, column=1).value or ""),
                        (ws.cell(row=r, column=2).value or ""),
                        (ws.cell(row=r, column=3).value or ""),
                        (ws.cell(row=r, column=4).value or ""),
                    ])
        except Exception as e:
            return {"ok": False, "error": f"Archivo inválido: {e}"}

        expected = ["depto", "telefono1", "telefono2", "nombre"]
        start_idx = 0
        if rows:
            header_lc = [str(x or "").strip().lower() for x in rows[0][:4]]
            if header_lc == expected:
                start_idx = 1

        processed_rows = 0
        for i in range(start_idx, len(rows)):
            try:
                c = rows[i]
                # Sanitiza celdas
                depto  = str(c[0] if len(c) > 0 else "").strip()
                tel1   = str(c[1] if len(c) > 1 else "").strip()
                tel2   = str(c[2] if len(c) > 2 else "").strip()
                nombre = str(c[3] if len(c) > 3 else "").strip()

                if not depto or not (tel1 or tel2):
                    skipped += 1
                    continue

                # Normaliza teléfonos
                tels: List[str] = []
                n1 = _norm_phone(tel1) if tel1 else ""
                n2 = _norm_phone(tel2) if tel2 else ""
                if n1: tels.append(n1)
                if n2 and n2 != n1: tels.append(n2)
                if not tels:
                    skipped += 1
                    continue

                # ----- UPSERT por (depto, nombre) -----
                key = {"departamento": depto, "nombre": (nombre or depto)}

                existing = self.col.find_one(key)
                if not existing:
                    telefono_field: Any = tels[0] if len(tels) == 1 else tels
                    doc = {
                        **key,
                        "telefono": telefono_field,
                        "createdAt": datetime.utcnow(),
                    }
                    self.col.insert_one(doc)
                    inserted += 1
                else:
                    # fusionar teléfonos
                    cur = existing.get("telefono")
                    cur_set: List[str]
                    if isinstance(cur, list):
                        cur_set = [ _norm_phone(str(x)) for x in cur if _norm_phone(str(x)) ]
                    elif cur:
                        cur_set = [ _norm_phone(str(cur)) ]
                    else:
                        cur_set = []

                    new_set = sorted(set(cur_set) | set(tels))
                    # ¿cambió algo?
                    if new_set != cur_set:
                        new_value: Any = new_set[0] if len(new_set) == 1 else new_set
                        self.col.update_one(key, {"$set": {"telefono": new_value, "_updatedAt": datetime.utcnow()}})
                        updated += 1
                    else:
                        skipped += 1

                processed_rows += 1

            except Exception as e:
                errors.append(f"Fila {i+1}: {e}")

        logging.info(f"Import CSV/XLSX => rows={processed_rows} inserted={inserted} updated={updated} skipped={skipped} errors={len(errors)}")
        return {"ok": True, "data": {
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "importedCount": inserted + updated,
            "errors": errors
        }}



    def _export_contacts(self, p: Dict[str, Any]) -> Dict[str, Any]:
        if Workbook is None:
            return {"ok": False, "error": "openpyxl no instalado en el contenedor"}

        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"

        # Cabeceras: depto - telefono 1 - telefono 2 - nombre
        ws.cell(row=1, column=1, value="depto")
        ws.cell(row=1, column=2, value="telefono 1")
        ws.cell(row=1, column=3, value="telefono 2")
        ws.cell(row=1, column=4, value="nombre")

        r = 2
        cur = self.col.find({}).sort([("departamento", ASCENDING), ("nombre", ASCENDING)])
        total = 0
        for d in cur:
            depto = d.get("departamento", "")
            nombre = d.get("nombre", "")
            tel = d.get("telefono", "")

            tel1, tel2 = "", ""
            if isinstance(tel, list):
                if len(tel) > 0: tel1 = _norm_phone(str(tel[0]))
                if len(tel) > 1: tel2 = _norm_phone(str(tel[1]))
            else:
                tel1 = _norm_phone(str(tel))

            ws.cell(row=r, column=1, value=depto)
            ws.cell(row=r, column=2, value=tel1)
            ws.cell(row=r, column=3, value=tel2)
            ws.cell(row=r, column=4, value=nombre)
            r += 1
            total += 1

        bio = BytesIO()
        wb.save(bio)
        b64 = base64.b64encode(bio.getvalue()).decode("ascii")
        return {"ok": True, "data": {"fileContent": b64, "count": total}}


def main():
    svc = ContactosService()
    try:
        if not svc.init_db():
            logging.error("No se pudo iniciar MongoDB")
            return
        if not svc.connect_bus():
            logging.error("No se pudo registrar en el BUS")
            return
        while svc.connected:
            time.sleep(1)
    except KeyboardInterrupt:
        logging.info("\nDeteniendo servicio...")
    finally:
        svc.disconnect()

if __name__ == "__main__":
    main()
